"""
Pipeline combinado y AUTOCONTENIDO: conectividad -> lectura PPPoE -> cruce con
contratos -> Excel final. No depende de prueba.py ni de ningún otro archivo
del toolchain (solo necesita DATOS.csv al lado).

Combina en un solo flujo lo que antes eran 3 scripts separados:
  1. check_conectividad.py  -> escanea el rango y detecta qué IPs responden en el
                               puerto de gestión (no hace login, no manda credenciales).
  2. obtener_pppoe.py       -> hace login (SOLO LECTURA) en cada IP online y lee el
                               Username PPPoE actual.
  3. cruzar_contratos.py    -> cruza ese PPPoE contra DATOS.csv para obtener tag,
                               onu_modelo, ativo_acs y status_contrato.

Y agrega la columna CAMBIAR_PPPOE (calculada en Python, no como fórmula de Excel):
  "No"  -> el pppoe ya termina en @wsneo.com.py (está correcto)
  "Sí"  -> el pppoe NO termina en @wsneo.com.py (hay que corregirlo)

Salida: un único archivo Excel (no se generan CSV intermedios).

IMPORTANTE: el login (paso 2) corre SECUENCIAL a propósito. Se reutiliza la
misma sesión de requests por router y se pasa la URL como parámetro (no como
variable global), así que técnicamente sería seguro paralelizarlo -- pero se
deja secuencial para no generar picos de carga sobre muchos routers a la vez.
"""

import csv
import hashlib
import hmac
import ipaddress
import re
import secrets
import socket
import time
import concurrent.futures

import requests
import openpyxl
from openpyxl.styles import Font

requests.packages.urllib3.disable_warnings()

# ============================= CONFIGURACIÓN =================================

RANGO = "100.90.66.0/24"          # <-- CIDR a escanear, ajustar según el lote
PORT = 1771
TIMEOUT_CONECTIVIDAD = 3          # segundos, para el escaneo de puerto
MAX_WORKERS_CONECTIVIDAD = 20     # hilos para el escaneo de conectividad (paralelo, seguro)

DELAY_ENTRE_ROUTERS = 2           # segundos, pausa entre logins (secuencial)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "N&0telec0m"

DOMINIO_ESPERADO = "@wsneo.com.py"

DATOS_CSV = "DATOS.csv"           # planilla de contratos (tag, pppoe, onu_modelo, ativo_acs, status_contrato)
OUTPUT_XLSX = "No_conecta_Migracion.xlsx"
HOJA = "Juan"

# ==============================================================================


# --------------------------- Login SCRAM (ex prueba.py) -----------------------

def _get_csrf(html_text: str) -> dict:
    param = re.search(r'name="csrf_param"\s+content="([^"]+)"', html_text)
    token = re.search(r'name="csrf_token"\s+content="([^"]+)"', html_text)
    if not param or not token:
        raise RuntimeError("No se encontraron csrf_param/csrf_token en el HTML")
    return {"csrf_param": param.group(1), "csrf_token": token.group(1)}


def _pbkdf2_sha256(password: str, salt: bytes, iterations: int, dklen: int = 32) -> bytes:
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations, dklen)


def _hmac_sha256(*, key: bytes, msg: bytes) -> bytes:
    return hmac.new(key, msg, hashlib.sha256).digest()


def _xor_bytes(a: bytes, b: bytes) -> bytes:
    return bytes(x ^ y for x, y in zip(a, b))


def login(session: requests.Session, base_url: str):
    """Login SCRAM contra el router. Devuelve (rsan, rsae, csrf)."""
    resp = session.get(f"{base_url}/html/index.html", verify=False, timeout=10)
    resp.raise_for_status()
    csrf = _get_csrf(resp.text)

    first_nonce = secrets.token_hex(32)
    payload_nonce = {
        "csrf": csrf,
        "data": {"username": ADMIN_USERNAME, "firstnonce": first_nonce},
    }
    resp = session.post(
        f"{base_url}/api/system/user_login_nonce",
        json=payload_nonce,
        headers={"Referer": f"{base_url}/html/index.html"},
        verify=False,
        timeout=10,
    )
    resp.raise_for_status()
    nonce_res = resp.json()

    if nonce_res.get("err") != 0:
        raise RuntimeError(f"user_login_nonce falló: {nonce_res}")

    salt = bytes.fromhex(nonce_res["salt"])
    iterations = int(nonce_res["iterations"])
    server_nonce = nonce_res["servernonce"]
    csrf = {"csrf_param": nonce_res["csrf_param"], "csrf_token": nonce_res["csrf_token"]}

    auth_msg = f"{first_nonce},{server_nonce},{server_nonce}"

    salted_password = _pbkdf2_sha256(ADMIN_PASSWORD, salt, iterations)
    client_key = _hmac_sha256(key=b"Client Key", msg=salted_password)
    stored_key = hashlib.sha256(client_key).digest()
    client_signature = _hmac_sha256(key=auth_msg.encode("utf-8"), msg=stored_key)
    client_proof = _xor_bytes(client_key, client_signature).hex()

    payload_proof = {
        "csrf": csrf,
        "data": {"clientproof": client_proof, "finalnonce": server_nonce},
    }

    resp = session.post(
        f"{base_url}/api/system/user_login_proof",
        json=payload_proof,
        headers={"Referer": f"{base_url}/html/index.html"},
        verify=False,
        timeout=10,
    )
    resp.raise_for_status()
    proof_res = resp.json()

    if proof_res.get("err") != 0:
        raise RuntimeError(f"user_login_proof falló (contraseña incorrecta?): {proof_res}")

    csrf = {"csrf_param": proof_res["csrf_param"], "csrf_token": proof_res["csrf_token"]}
    rsan = proof_res["rsan"]
    rsae = proof_res["rsae"]

    return rsan, rsae, csrf

# --------------------------------------------------------------------------


def escanear_conectividad(rango: str) -> list[str]:
    """Paso 1: qué IPs del rango responden en el puerto de gestión."""
    ip_list = [str(ip) for ip in ipaddress.ip_network(rango).hosts()]
    print(f"[1/3] Verificando conectividad al puerto {PORT} en {len(ip_list)} IP(s)...")

    def check_port(ip: str):
        try:
            with socket.create_connection((ip, PORT), timeout=TIMEOUT_CONECTIVIDAD):
                return ip, True
        except (socket.timeout, ConnectionRefusedError, OSError):
            return ip, False

    online = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS_CONECTIVIDAD) as executor:
        for ip, is_open in executor.map(check_port, ip_list):
            if is_open:
                online.append(ip)

    print(f"      {len(online)} de {len(ip_list)} IPs online.\n")
    return online


def leer_pppoe(base_url: str) -> dict:
    """Login (solo lectura) + lectura del Username PPPoE actual de un router."""
    session = requests.Session()
    session.headers.update({
        "Origin": base_url,
        "X-Requested-With": "XMLHttpRequest",
        "_responseformat": "JSON",
    })

    try:
        login(session, base_url)
        resp = session.get(f"{base_url}/api/ntwk/wan?type=active", verify=False, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        return {"url": base_url, "pppoe_actual": data.get("Username", ""), "detalle": ""}
    except requests.exceptions.ConnectionError as e:
        return {"url": base_url, "pppoe_actual": "", "detalle": f"Conexión: {e}"}
    except requests.exceptions.Timeout as e:
        return {"url": base_url, "pppoe_actual": "", "detalle": f"Timeout: {e}"}
    except RuntimeError as e:
        return {"url": base_url, "pppoe_actual": "", "detalle": f"Login: {e}"}
    except Exception as e:
        return {"url": base_url, "pppoe_actual": "", "detalle": f"Inesperado: {e}"}


def obtener_pppoe_de_todos(ips_online: list[str]) -> list[dict]:
    """Paso 2: PPPoE actual de cada IP online (secuencial)."""
    urls = [f"https://{ip}:{PORT}" for ip in ips_online]
    total = len(urls)
    print(f"[2/3] Leyendo PPPoE de {total} router(es)...")

    resultados = []
    for i, url in enumerate(urls, start=1):
        r = leer_pppoe(url)
        resultados.append(r)
        estado = r["pppoe_actual"] or f"ERROR: {r['detalle']}"
        print(f"      [{i}/{total}] {url} -> {estado}")
        if i < total:
            time.sleep(DELAY_ENTRE_ROUTERS)

    ok = sum(1 for r in resultados if r["pppoe_actual"])
    print(f"      Leídos correctamente: {ok} | Con error: {total - ok}\n")
    return resultados


def cargar_datos_contratos(path: str) -> dict:
    indice = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pppoe = (row.get("pppoe") or "").strip()
            if pppoe:
                indice[pppoe.lower()] = row
    return indice


def cruzar_y_armar_filas(resultados_pppoe: list[dict], indice_contratos: dict) -> list[dict]:
    """Paso 3: cruce contra DATOS.csv + cálculo de CAMBIAR_PPPOE."""
    print(f"[3/3] Cruzando {len(resultados_pppoe)} router(es) contra {DATOS_CSV}...")

    filas = []
    con_match = 0
    for r in resultados_pppoe:
        url = r["url"]
        pppoe_actual = (r.get("pppoe_actual") or "").strip()

        if not pppoe_actual:
            fila_contrato = None
            match = "SIN_PPPOE"
        else:
            fila_contrato = indice_contratos.get(pppoe_actual.lower())
            match = "SI" if fila_contrato else "NO"

        if fila_contrato:
            con_match += 1

        cambiar_pppoe = "No" if pppoe_actual.endswith(DOMINIO_ESPERADO) else "Sí"

        filas.append({
            "tag": fila_contrato.get("tag", "") if fila_contrato else "",
            "pppoe": pppoe_actual,
            "ip": url,
            "match": match,
            "onu_modelo": fila_contrato.get("onu_modelo", "") if fila_contrato else "",
            "ativo_acs": fila_contrato.get("ativo_acs", "") if fila_contrato else "",
            "status_contrato": fila_contrato.get("status_contrato", "") if fila_contrato else "",
            "CAMBIAR_PPPOE": cambiar_pppoe,
        })

    print(f"      Con match: {con_match} | Sin match: {len(filas) - con_match}\n")
    return filas


def escribir_excel(filas: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = HOJA

    headers = ["tag", "pppoe", "ip", "match", "onu_modelo", "ativo_acs", "status_contrato", "CAMBIAR_PPPOE"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for fila in filas:
        ws.append([fila[h] for h in headers])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    # Ancho de columnas aproximado
    anchos = {"A": 10, "B": 32, "C": 26, "D": 10, "E": 12, "F": 10, "G": 16, "H": 14}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho

    wb.save(path)
    print(f"[OK] Excel guardado en {path} ({len(filas)} filas)")


def main():
    ips_online = escanear_conectividad(RANGO)
    if not ips_online:
        print("No hay IPs online, no hay nada para procesar.")
        return

    resultados_pppoe = obtener_pppoe_de_todos(ips_online)
    indice_contratos = cargar_datos_contratos(DATOS_CSV)
    filas = cruzar_y_armar_filas(resultados_pppoe, indice_contratos)
    escribir_excel(filas, OUTPUT_XLSX)


if __name__ == "__main__":
    main()